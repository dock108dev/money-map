from __future__ import annotations

import csv
from collections.abc import Iterable
from datetime import date
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from ..money import money
from .base import Evidence, ParsedLedger, ParsedRecord, UnsupportedLayoutError

REQUIRED_COLUMNS = {
    "institution",
    "account",
    "date",
    "record_type",
    "role",
    "amount",
}

INSTITUTION_ALIASES = {
    "axos": "SoFi",
    "provident": "SoFi",
    "sofi": "SoFi",
    "fidelity": "Fidelity",
}

SOFI_ROLES = {
    "external_inflow",
    "external_outflow",
    "internal_transfer",
    "interest",
    "fee",
    "adjustment",
    "unresolved",
    "payroll_deposit",
}

FIDELITY_ROLES = {
    "employee_contribution",
    "employer_contribution",
    "stock_plan_contribution",
    "external_deposit",
    "external_withdrawal",
    "internal_transfer",
    "purchase",
    "sale",
    "dividend",
    "interest",
    "reinvestment",
    "fee",
    "adjustment",
    "unresolved",
}


def canonical_institution(raw: str) -> str:
    lowered = raw.strip().lower()
    return INSTITUTION_ALIASES.get(lowered, raw.strip())


def _account_type(institution: str, account: str) -> str:
    lowered = account.lower()
    if institution == "SoFi":
        return "savings" if "sav" in lowered else "checking"
    if institution == "Fidelity":
        return "investment"
    return "other"


def _date(raw: Any) -> date:
    if isinstance(raw, date):
        return raw
    return date.fromisoformat(str(raw).strip())


def _optional_money(raw: Any) -> Decimal | None:
    if raw in (None, ""):
        return None
    return money(str(raw).replace(",", "").replace("$", ""))


class CanonicalLedgerAdapter:
    """Portable manual CSV/XLSX contract for SoFi and Fidelity records."""

    name = "canonical_ledger"
    parser_version = "1.0.0"

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in {".csv", ".xlsx"}

    def parse(self, path: Path) -> ParsedLedger:
        rows = list(self._rows(path))
        if not rows:
            raise ValueError("Ledger file is empty")
        normalized_headers = {str(key).strip().lower() for key in rows[0]}
        if not REQUIRED_COLUMNS.issubset(normalized_headers):
            missing = sorted(REQUIRED_COLUMNS - normalized_headers)
            raise UnsupportedLayoutError(f"Missing canonical ledger columns: {missing}")

        records: list[ParsedRecord] = []
        for row_number, source in enumerate(rows, start=2):
            row = {str(key).strip().lower(): value for key, value in source.items()}
            institution = canonical_institution(str(row["institution"]))
            role = str(row["role"]).strip().lower()
            record_type = str(row["record_type"]).strip().lower()
            if record_type not in {"transaction", "balance"}:
                raise ValueError(f"Row {row_number}: record_type must be transaction or balance")
            if record_type == "balance" and role not in {"opening", "closing"}:
                raise ValueError(f"Row {row_number}: balance role must be opening or closing")
            allowed_roles = SOFI_ROLES if institution == "SoFi" else FIDELITY_ROLES
            if record_type == "transaction" and role not in allowed_roles:
                role = "unresolved"
            account = str(row["account"]).strip()
            amount_value = _optional_money(row["amount"])
            if amount_value is None:
                raise ValueError(f"Row {row_number}: amount is required")
            records.append(
                ParsedRecord(
                    institution=institution,
                    account=account,
                    account_type=_account_type(institution, account),
                    record_date=_date(row["date"]),
                    record_type=record_type,
                    role=role,
                    amount=amount_value,
                    description=str(row.get("description") or "").strip(),
                    balance=_optional_money(row.get("balance")),
                    row_number=row_number,
                )
            )
        return ParsedLedger(
            records=tuple(records),
            evidence=(
                Evidence(
                    "ledger_rows",
                    "rows 2 onward",
                    "canonical ledger columns",
                    extraction_method=path.suffix.lower().lstrip("."),
                ),
            ),
        )

    def _rows(self, path: Path) -> Iterable[dict[str, Any]]:
        if path.suffix.lower() == ".csv":
            with path.open(newline="", encoding="utf-8-sig") as handle:
                yield from csv.DictReader(handle)
            return
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Workbook has no active worksheet")
        iterator = sheet.iter_rows(values_only=True)
        headers = next(iterator, None)
        if headers is None:
            return
        for values in iterator:
            yield dict(zip((str(header) for header in headers), values, strict=True))
